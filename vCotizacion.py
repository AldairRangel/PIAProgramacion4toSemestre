from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import math
from openpyxl import load_workbook
from datetime import date
from win32com import client
import win32api
import pathlib

class Ui_cotizacion(object):
    def setupUi(self, Cotizacion):
        Cotizacion.setObjectName("Cotizacion")
        Cotizacion.resize(558, 497)
        self.centralwidget = QtWidgets.QWidget(Cotizacion)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(-10, 0, 581, 501))
        self.label.setStyleSheet("color: rgb(0, 0, 0);\n"
"background-color: rgb(0, 0, 0);")
        self.label.setText("")
        self.label.setObjectName("label")
        self.bim4 = QtWidgets.QTextEdit(self.centralwidget)
        self.bim4.setGeometry(QtCore.QRect(150, 190, 51, 41))
        self.bim4.setStyleSheet("font: 12pt \"MS Shell Dlg 2\";")
        self.bim4.setObjectName("bim4")
        self.bim5 = QtWidgets.QTextEdit(self.centralwidget)
        self.bim5.setGeometry(QtCore.QRect(150, 240, 51, 41))
        self.bim5.setStyleSheet("font: 12pt \"MS Shell Dlg 2\";")
        self.bim5.setObjectName("bim5")
        self.bim6 = QtWidgets.QTextEdit(self.centralwidget)
        self.bim6.setGeometry(QtCore.QRect(150, 290, 51, 41))
        self.bim6.setStyleSheet("font: 12pt \"MS Shell Dlg 2\";")
        self.bim6.setObjectName("bim6")
        self.bim1 = QtWidgets.QTextEdit(self.centralwidget)
        self.bim1.setGeometry(QtCore.QRect(150, 40, 51, 41))
        self.bim1.setStyleSheet("font: 12pt \"MS Shell Dlg 2\";")
        self.bim1.setObjectName("bim1")
        self.bim2 = QtWidgets.QTextEdit(self.centralwidget)
        self.bim2.setGeometry(QtCore.QRect(150, 90, 51, 41))
        self.bim2.setStyleSheet("font: 12pt \"MS Shell Dlg 2\";")
        self.bim2.setObjectName("bim2")
        self.bim3 = QtWidgets.QTextEdit(self.centralwidget)
        self.bim3.setGeometry(QtCore.QRect(150, 140, 51, 41))
        self.bim3.setStyleSheet("font: 12pt \"MS Shell Dlg 2\";")
        self.bim3.setObjectName("bim3")
        self.pb_calcular = QtWidgets.QPushButton(self.centralwidget)
        self.pb_calcular.setGeometry(QtCore.QRect(270, 230, 75, 23))
        self.pb_calcular.setObjectName("pb_calcular")
        self.khwdiarios = QtWidgets.QTextBrowser(self.centralwidget)
        self.khwdiarios.setGeometry(QtCore.QRect(300, 280, 201, 61))
        self.khwdiarios.setStyleSheet("font: 11pt \"MS Shell Dlg 2\";")
        self.khwdiarios.setObjectName("khwdiarios")
        self.pb_borrardatos = QtWidgets.QPushButton(self.centralwidget)
        self.pb_borrardatos.setGeometry(QtCore.QRect(450, 230, 75, 23))
        self.pb_borrardatos.setObjectName("pb_borrardatos")
        self.totalpaneles = QtWidgets.QTextBrowser(self.centralwidget)
        self.totalpaneles.setGeometry(QtCore.QRect(300, 350, 201, 51))
        self.totalpaneles.setStyleSheet("font: 11pt \"MS Shell Dlg 2\";")
        self.totalpaneles.setObjectName("totalpaneles")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(340, 420, 201, 31))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(70, 50, 81, 21))
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(70, 100, 81, 21))
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(70, 150, 81, 21))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(70, 200, 81, 21))
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(70, 250, 81, 21))
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        self.label_8.setGeometry(QtCore.QRect(70, 300, 81, 21))
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.centralwidget)
        self.label_9.setGeometry(QtCore.QRect(10, 370, 81, 21))
        self.label_9.setObjectName("label_9")
        self.sumakhw = QtWidgets.QTextBrowser(self.centralwidget)
        self.sumakhw.setGeometry(QtCore.QRect(90, 350, 131, 51))
        self.sumakhw.setObjectName("sumakhw")
        self.pb_generarpdf = QtWidgets.QPushButton(self.centralwidget)
        self.pb_generarpdf.setGeometry(QtCore.QRect(360, 230, 75, 23))
        self.pb_generarpdf.setObjectName("pb_generarpdf")
        self.nombre_cliente = QtWidgets.QTextEdit(self.centralwidget)
        self.nombre_cliente.setGeometry(QtCore.QRect(280, 80, 251, 41))
        self.nombre_cliente.setStyleSheet("font: 11pt \"MS Shell Dlg 2\";")
        self.nombre_cliente.setObjectName("nombre_cliente")
        self.label_11 = QtWidgets.QLabel(self.centralwidget)
        self.label_11.setGeometry(QtCore.QRect(280, 120, 251, 21))
        self.label_11.setObjectName("label_11")
        self.direccion_cliente = QtWidgets.QTextEdit(self.centralwidget)
        self.direccion_cliente.setGeometry(QtCore.QRect(280, 140, 251, 41))
        self.direccion_cliente.setStyleSheet("font: 11pt \"MS Shell Dlg 2\";")
        self.direccion_cliente.setObjectName("direccion_cliente")
        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        self.label_10.setGeometry(QtCore.QRect(280, 60, 251, 21))
        self.label_10.setObjectName("label_10")
        Cotizacion.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Cotizacion)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 558, 21))
        self.menubar.setObjectName("menubar")
        Cotizacion.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Cotizacion)
        self.statusbar.setObjectName("statusbar")
        Cotizacion.setStatusBar(self.statusbar)

        self.retranslateUi(Cotizacion)
        self.pb_borrardatos.clicked.connect(self.bim1.clear)
        self.pb_borrardatos.clicked.connect(self.bim2.clear)
        self.pb_borrardatos.clicked.connect(self.bim3.clear)
        self.pb_borrardatos.clicked.connect(self.bim4.clear)
        self.pb_borrardatos.clicked.connect(self.bim5.clear)
        self.pb_borrardatos.clicked.connect(self.bim6.clear)
        self.pb_borrardatos.clicked.connect(self.khwdiarios.clear)
        self.pb_borrardatos.clicked.connect(self.totalpaneles.clear)
        self.pb_borrardatos.clicked.connect(self.sumakhw.clear)
        QtCore.QMetaObject.connectSlotsByName(Cotizacion)

    def retranslateUi(self, Cotizacion):
        _translate = QtCore.QCoreApplication.translate
        Cotizacion.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.pb_calcular.setText(_translate("MainWindow", "Calcular"))
        self.pb_borrardatos.setText(_translate("MainWindow", "Borrar datos"))
        self.label_2.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:11pt; font-weight:600; color:#ffff00;\">monterreypanelsolar.com</span></p></body></html>"))
        self.label_3.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Bimestre 1</span></p></body></html>"))
        self.label_4.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Bimestre 2</span></p></body></html>"))
        self.label_5.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Bimestre 3</span></p></body></html>"))
        self.label_6.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Bimestre 4</span></p></body></html>"))
        self.label_7.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Bimestre 5</span></p></body></html>"))
        self.label_8.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Bimestre 6</span></p></body></html>"))
        self.label_9.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color:#ffffff;\">Suma anual</span></p></body></html>"))
        self.pb_generarpdf.setText(_translate("MainWindow", "Generar PDF"))
        self.label_11.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Ingresa dirección:</span></p></body></html>"))
        self.label_10.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Ingresa nombre del cliente:</span></p></body></html>"))
        self.pb_calcular.clicked.connect(self.calculopaneles)
        self.pb_generarpdf.clicked.connect(self.generarpdf)

    def calculopaneles(self):
        B1 = int(self.bim1.toPlainText())        
        B2 = int(self.bim2.toPlainText())
        B3 = int(self.bim3.toPlainText())
        B4 = int(self.bim4.toPlainText())
        B5 = int(self.bim5.toPlainText())
        B6 = int(self.bim6.toPlainText())
        
        sumabim = B1+B2+B3+B4+B5+B6
        mensaje_sumaanual = f"El consumo de khw anual que presenta es de: " + str(sumabim)
        khwxdia = sumabim/6/60
        mensaje_khwdiarios = f"El consumo promedio diario de khw es: " + str(khwxdia)
        
        paneles_necesarios = khwxdia/1.7
        redondear_punto5 = round(paneles_necesarios)
        mensaje_panelesnecesarios = f"La cantidad de paneles requerida es de: " + str(redondear_punto5)

        NombreCliente = str(self.nombre_cliente.toPlainText())
        DireccionCliente = str(self.direccion_cliente.toPlainText())
        FechaEnExcel = date.today()

        filename = "CotizacionPANELES.xlsx"

        wb = load_workbook(filename)
        ws = wb.worksheets[0]

        ws_tables = []
        ws["B1"] = " " + str(NombreCliente)
        ws["B2"] = " " + str(DireccionCliente)
        ws["B4"] = " " + str(sumabim)
        ws["B6"] = " " + str(date.today())
        ws["C13"] = " " + str(redondear_punto5)
        
        for table in ws._tables:
            ws_tables.append(table)
        wb.save(filename)

        self.sumakhw.setText(mensaje_sumaanual)
        self.khwdiarios.setText(mensaje_khwdiarios)
        self.totalpaneles.setText(mensaje_panelesnecesarios)

    def generarpdf(self):
        excel_file = "CotizacionPANELES.xlsx"
        pdf_file = "Cotizacion.pdf"
        excel_path = str(pathlib.Path.cwd() / excel_file)
        pdf_path = str(pathlib.Path.cwd() / pdf_file)
        excel = client.DispatchEx("Excel.Application")
        excel.Visible = 0
        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Worksheets[1]

        try:
            wb.SaveAs(pdf_path, FileFormat=57)
        except Exception as e:
            print("Failed to convert")
            print(str(e))
        finally:
            wb.Close()
            excel.Quit()

if __name__ == "__main__":
    #import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
